# Developed by abaye ©
# 💉 Converts years to Hebrew v1.0.0 19/05/2024			
# https://ko-fi.com/abaye
# email: cs@abaye.co

import os
import sys
import openpyxl
import datetime
from convertdate import hebrew
#import hdate
import re


def get_script_directory():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))


def hebrew_number_to_str(year):
    hebrew_letters = {
        1: 'א', 2: 'ב', 3: 'ג', 4: 'ד', 5: 'ה', 6: 'ו', 7: 'ז', 8: 'ח', 9: 'ט',
        10: 'י', 20: 'כ', 30: 'ל', 40: 'מ', 50: 'נ', 60: 'ס', 70: 'ע', 80: 'פ', 90: 'צ',
        100: 'ק', 200: 'ר', 300: 'ש', 400: 'ת'
    }

    result = ""
    
    # טיפול באלפים
    if year >= 1000:
        thousands = year // 1000
        if thousands > 1:
            result += hebrew_letters[thousands]
        result += "' "
        year %= 1000

    # טיפול במאות גדולות מ400
    while year >= 400:
        result += hebrew_letters[400]
        year -= 400

    # מציאת מאות
    if year >= 100:
        hundreds = year // 100 * 100
        result += hebrew_letters[hundreds]
        year -= hundreds

    # מציאת עשרות
    if year >= 10:
        tens = year // 10 * 10
        result += hebrew_letters[tens]
        year -= tens

    # מציאת יחידות
    if year > 0:
        result += hebrew_letters[year]

    return result


def convert_gregorian_to_hebrew(year):
    hebrew_date = hebrew.from_gregorian(year, 1, 1)
    hebrew_year = hebrew_date[0]

    #gregorian_date = datetime.date(year, 1, 1)
    #hebrew_date = hdate.HDate(gregorian_date, hebrew=True)
 
    hebrew_year_str = hebrew_number_to_str(hebrew_year)
    return f'{hebrew_year_str}'


def extract_numbers(value):
    numbers = re.findall(r'\d+', value)

    if len(numbers) == 1:
        hebrew_year = convert_gregorian_to_hebrew(int(numbers[0]))
        print(hebrew_year)
        return f'{hebrew_year}'
    elif len(numbers) == 2:
        hebrew_year_start = convert_gregorian_to_hebrew(int(numbers[0]))
        hebrew_year_end = convert_gregorian_to_hebrew(int(numbers[0]))
        print(f'{hebrew_year_start} - {hebrew_year_end}')
        return f'{hebrew_year_start} - {hebrew_year_end}'


def file_process():
    workbook = openpyxl.load_workbook(metadata_file_path)
    sheet = workbook.active

    for row in range(2, sheet.max_row + 1):
        for col in [3, 6]:
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None:
                text = cell.value
                print(text)
                new_value = extract_numbers(text)

                if "בקירוב" in text:
                    cell.value = f'{new_value} (בקירוב)'
                else:
                    cell.value = new_value

    workbook.save(metadata_convert_file_path)



script_directory = get_script_directory()
metadata_file_path = f'{script_directory}\metadata.xlsx'
metadata_convert_file_path = f'{script_directory}\metadata_convert_he.xlsx'

file_process()